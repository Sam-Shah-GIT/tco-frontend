# ui/utils.py
from openpyxl import load_workbook
from django.contrib import messages

def read_model_options_from_excel(file_path, sheet_name='Sheet1'):
    try:
        wb = load_workbook(filename=file_path)
        ws = wb[sheet_name]
        unique_models = [ws.cell(row=i, column=1).value for i in range(6, ws.max_row + 1)]  # Assuming data starts from row 6
        model_options = [model for model in unique_models if model is not None]
        return model_options
    except FileNotFoundError:
        return []
    except Exception as e:
        return []
